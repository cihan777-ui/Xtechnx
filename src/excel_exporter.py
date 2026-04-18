"""
Platform Excel Şablon Doldurucu
- Hepsiburada: 5 sütunlu Hızlı Ürün Ekleme şablonu
- N11: 32 sütunlu Toplu Ürün Ekle şablonu
"""
import random
import string
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPORTS_DIR = Path("exports")

# Renkler
C_HEADER_HB  = "FF6000"   # Hepsiburada turuncu
C_HEADER_N11 = "4A90D9"   # N11 mavi
C_ROW_ALT    = "F8F9FF"
C_REQUIRED   = "FFF3CD"   # Zorunlu alan sarısı
C_WHITE      = "FFFFFF"


def _thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


# ════════════════════════════════════════════════════════
#  HEPSİBURADA
# ════════════════════════════════════════════════════════

HB_COLUMNS = ["Ürün Adı", "Satıcı Stok Kodu", "Barkod", "Fiyat", "Stok"]

def generate_hepsiburada(products: list) -> str:
    """
    products: [{"title","sku","barcode","price","stock"}, ...]
    Hepsiburada Hızlı Ürün Ekleme şablonunu doldurur.
    """
    EXPORTS_DIR.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Fast Listing"
    ws.sheet_view.showGridLines = False

    # Satır 1: Zorunlu işaretleri
    for i in range(1, 4):
        cell = ws.cell(1, i, "Zorunlu")
        cell.font = Font(bold=True, color="FF0000", size=9)
        cell.alignment = Alignment(horizontal="center")

    # Satır 2: Başlıklar
    header_fill = PatternFill("solid", fgColor=C_HEADER_HB)
    for i, col in enumerate(HB_COLUMNS, 1):
        cell = ws.cell(2, i, col)
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 22

    # Satır 3+: Ürünler
    for r, p in enumerate(products, 3):
        alt = r % 2 == 0
        row_fill = PatternFill("solid", fgColor=C_ROW_ALT) if alt else None
        values = [
            p.get("title", ""),
            p.get("sku", ""),
            p.get("barcode", ""),
            p.get("price", 0),
            p.get("stock", 10),
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(r, c, val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = _thin_border()
            if row_fill:
                cell.fill = row_fill
            if c == 4:  # Fiyat
                cell.number_format = '#,##0.00'
            if c == 5:  # Stok
                cell.number_format = '#,##0'

    # Sütun genişlikleri
    widths = [55, 22, 18, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    fname = EXPORTS_DIR / f"hepsiburada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(fname))
    return str(fname)


# ════════════════════════════════════════════════════════
#  N11
# ════════════════════════════════════════════════════════

N11_COLUMNS = [
    "Stok Kodu", "Model Kodu ", "Marka ", "Kategori ", "Para Birimi",
    "Ürün Adı", "Ürün Açıklaması",
    "Piyasa Satış Fiyatı (KDV Dahil)", "N11 Satış Fiyatı (KDV Dahil)",
    "Stok", "KDV Oranı",
    "Görsel 1", "Görsel 2", "Görsel 3", "Görsel 4", "Görsel 5",
    "Görsel 6", "Görsel 7", "Görsel 8", "Görsel 9", "Görsel 10",
    "Görsel 11", "Görsel 12",
    "Hazırlık Süresi ", "Teslimat Şablonu İsmi", "Katalog ID",
    "Barkod (GTIN,EAN)", "Maksimum Satış Adedi",
    "Giriş Gerilimi (Düşük)", "Kapasite (VA)", "Ses Seviyesi", "Ups Türü",
]

def generate_n11(products: list, category_id: int = 0,
                 delivery_template: str = "Standart Teslimat") -> str:
    """
    products: [{"title","description","sku","barcode","price","stock",
                "brand","images":[], "category_id"(opsiyonel)}, ...]
    N11 Toplu Ürün Ekle şablonunu doldurur.
    """
    EXPORTS_DIR.mkdir(exist_ok=True)

    # Orijinal şablonu temel alarak yükle (varsa)
    template_path = Path("templates/n11_template.xlsx")
    if template_path.exists():
        wb = load_workbook(str(template_path))
        ws = wb["Toplu Ürün Ekle"]
        # Mevcut veri satırlarını temizle (3. satırdan itibaren)
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.value = None
        start_row = 3
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Toplu Ürün Ekle"
        ws.sheet_view.showGridLines = False

        # Başlık satırı
        header_fill = PatternFill("solid", fgColor=C_HEADER_N11)
        for i, col in enumerate(N11_COLUMNS, 1):
            cell = ws.cell(1, i, col)
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = _thin_border()
        ws.row_dimensions[1].height = 30
        start_row = 2

    # Ürün satırları
    for r, p in enumerate(products, start_row):
        alt = r % 2 == 0
        row_fill = PatternFill("solid", fgColor=C_ROW_ALT) if alt else None

        images = p.get("images", [])
        # 12 görsel slotu
        img_values = images[:12] + [""] * (12 - len(images[:12]))

        cat_id = p.get("category_id") or category_id or 0

        # Piyasa fiyatı = satış fiyatının %20 üstü (öneri)
        sale_price   = float(p.get("price", 0))
        market_price = round(sale_price * 1.20, 2)

        row_values = [
            p.get("sku", ""),           # 1. Stok Kodu
            p.get("sku", ""),           # 2. Model Kodu
            p.get("brand", "Xtechnx"),  # 3. Marka
            cat_id,                     # 4. Kategori
            "TL",                       # 5. Para Birimi
            p.get("title", ""),         # 6. Ürün Adı
            p.get("description", ""),   # 7. Ürün Açıklaması
            market_price,               # 8. Piyasa Fiyatı
            sale_price,                 # 9. N11 Fiyatı
            p.get("stock", 10),         # 10. Stok
            18,                         # 11. KDV Oranı
            *img_values,                # 12-23. Görseller
            3,                          # 24. Hazırlık Süresi
            delivery_template,          # 25. Teslimat Şablonu
            "",                         # 26. Katalog ID
            p.get("barcode", ""),       # 27. Barkod
            100,                        # 28. Maks. Satış Adedi
            "", "", "", "",             # 29-32. Kategori özellikleri (UPS)
        ]

        for c, val in enumerate(row_values, 1):
            cell = ws.cell(r, c, val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 7))
            cell.border = _thin_border()
            if row_fill:
                cell.fill = row_fill
            if c in (8, 9):  # Fiyatlar
                cell.number_format = '#,##0.00'
            if c == 10:       # Stok
                cell.number_format = '#,##0'

    # Sütun genişlikleri
    col_widths = [
        18, 18, 15, 12, 10,   # 1-5
        50, 40,               # 6-7 (başlık, açıklama)
        16, 16, 10, 10,       # 8-11
        35, 35, 35, 35, 35,   # 12-16 görseller
        35, 35, 35, 35, 35,   # 17-21 görseller
        35, 35,               # 22-23 görseller
        14, 25, 12,           # 24-26
        18, 12,               # 27-28
        15, 15, 15, 15,       # 29-32
    ]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    fname = EXPORTS_DIR / f"n11_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(fname))
    return str(fname)


# ════════════════════════════════════════════════════════
#  TOPLU ÜRETME
# ════════════════════════════════════════════════════════

def generate_all(products: list, platforms: list = None,
                 n11_category_id: int = 0) -> dict:
    """
    Seçili platformlar için tüm Excel dosyalarını üretir.
    Döner: {"hepsiburada": "exports/hb_xxx.xlsx", "n11": "exports/n11_xxx.xlsx"}
    """
    if platforms is None:
        platforms = ["hepsiburada", "n11"]

    results = {}
    for platform in platforms:
        try:
            if platform == "hepsiburada":
                results["hepsiburada"] = generate_hepsiburada(products)
            elif platform == "n11":
                results["n11"] = generate_n11(products, category_id=n11_category_id)
        except Exception as e:
            results[platform] = {"error": str(e)}

    return results
