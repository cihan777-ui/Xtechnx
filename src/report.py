"""
Excel Rapor Üreticisi
Yükleme geçmişini .xlsx formatında dışa aktarır.
"""
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database as db


REPORTS_DIR = Path("reports")

# Renkler
COLOR_HEADER   = "1a1a2e"
COLOR_SUCCESS  = "d4edda"
COLOR_ERROR    = "f8d7da"
COLOR_ALT_ROW  = "f8f9ff"
COLOR_ACCENT   = "7c5cfc"


def _header_style(ws, row, cols):
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        thin = Side(style="thin", color="FFFFFF")
        cell.border = Border(bottom=thin)


def generate_history_report(platform=None, status=None) -> str:
    """Yükleme geçmişi raporu oluşturur. Dosya yolunu döner."""
    REPORTS_DIR.mkdir(exist_ok=True)

    records = db.get_history(limit=10000, platform=platform, status=status)
    stats = db.get_history_stats()

    wb = openpyxl.Workbook()

    # ── Özet Sayfası ────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Özet"
    ws_sum.sheet_view.showGridLines = False

    ws_sum["A1"] = "Xtechnx Product Sync — Yükleme Raporu"
    ws_sum["A1"].font = Font(bold=True, size=16, color=COLOR_ACCENT)
    ws_sum["A2"] = f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_sum["A2"].font = Font(size=10, color="888888")

    ws_sum.row_dimensions[1].height = 30
    ws_sum.row_dimensions[3].height = 8

    headers_sum = ["Platform", "Toplam", "Başarılı", "Hatalı"]
    for i, h in enumerate(headers_sum, 1):
        ws_sum.cell(4, i, h)
    _header_style(ws_sum, 4, len(headers_sum))

    platforms = ['trendyol', 'hepsiburada', 'n11', 'amazon']
    platform_labels = {
        'trendyol': 'Trendyol', 'hepsiburada': 'Hepsiburada',
        'n11': 'N11', 'amazon': 'Amazon TR'
    }
    for r, plat in enumerate(platforms, 5):
        s = stats.get(plat, {})
        ws_sum.cell(r, 1, platform_labels.get(plat, plat))
        ws_sum.cell(r, 2, s.get('total', 0))
        ws_sum.cell(r, 3, s.get('success', 0)).font = Font(color="1a7a4a", bold=True)
        ws_sum.cell(r, 4, s.get('error', 0)).font = Font(color="a94442", bold=True)
        if r % 2 == 0:
            for c in range(1, 5):
                ws_sum.cell(r, c).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

    ws_sum.cell(10, 1, "Toplam Yükleme").font = Font(bold=True)
    ws_sum.cell(10, 2, stats.get('total', 0)).font = Font(bold=True)
    ws_sum.cell(11, 1, "Bugün").font = Font(bold=True)
    ws_sum.cell(11, 2, stats.get('today', 0)).font = Font(bold=True)

    for i, w in enumerate([20, 12, 12, 12], 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    # ── Detay Sayfası ────────────────────────────────────────
    ws_det = wb.create_sheet("Detay")
    ws_det.sheet_view.showGridLines = False

    headers_det = [
        "Tarih", "Platform", "Durum",
        "Orijinal Barkod", "Yeni Barkod (Xtechnx)", "Stok Kodu (Cihan)",
        "Orijinal Başlık", "Yüklenen Başlık",
        "Orijinal Fiyat", "Satış Fiyatı",
        "Hata Mesajı"
    ]
    col_widths = [18, 12, 10, 18, 22, 18, 35, 35, 14, 14, 40]

    for i, h in enumerate(headers_det, 1):
        ws_det.cell(1, i, h)
    _header_style(ws_det, 1, len(headers_det))
    ws_det.row_dimensions[1].height = 22
    ws_det.freeze_panes = "A2"

    for r, rec in enumerate(records, 2):
        ok = rec['status'] == 'success'
        row_fill = PatternFill("solid", fgColor=COLOR_SUCCESS if ok else COLOR_ERROR) if not ok else \
                   (PatternFill("solid", fgColor=COLOR_ALT_ROW) if r % 2 == 0 else None)

        values = [
            rec['uploaded_at'][:16].replace('T', ' '),
            platform_labels.get(rec['platform'], rec['platform']),
            '✓ Başarılı' if ok else '✗ Hata',
            rec['barcode_orig'],
            rec['barcode_new'],
            rec['sku_new'],
            rec['title_orig'],
            rec['title_new'],
            rec['price_orig'],
            rec['price_new'],
            rec['error_msg'] or '',
        ]
        for c, val in enumerate(values, 1):
            cell = ws_det.cell(r, c, val)
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill
            if c == 3:
                cell.font = Font(color="1a7a4a" if ok else "a94442", bold=True)

    for i, w in enumerate(col_widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w

    # ── Kaydet ──────────────────────────────────────────────
    fname = REPORTS_DIR / f"xtechnx_rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(fname))
    return str(fname)
